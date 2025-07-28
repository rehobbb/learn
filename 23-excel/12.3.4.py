import openpyxl  # 导入 openpyxl 库，用于操作 Excel 文件
from openpyxl.utils import get_column_letter, column_index_from_string  # 从 openpyxl.utils 模块导入获取列字母和列字母转索引的函数
print(get_column_letter(1))  # 打印第 1 列对应的列字母
print(get_column_letter(27))  # 打印第 27 列对应的列字母
wb = openpyxl.load_workbook('./doc/example.xlsx')  # 加载指定路径下的 Excel 文件
sheet = wb['Sheet1']  # 获取名为 'Sheet1' 的工作表
print(get_column_letter(sheet.max_column))  # 打印工作表最大列数对应的列字母
number = column_index_from_string('A')  # 将列字母 'A' 转换为对应的列索引
print(number)  # 打印列字母 'A' 对应的列索引
