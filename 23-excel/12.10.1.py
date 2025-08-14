import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['a1'] = 'hello world'
sheet['b2'] = 'nihao shenzhen'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
sheet.merge_cells('a1:b2')
sheet.unmerge_cells('a1:b2')
wb.save('dimension.xlsx')