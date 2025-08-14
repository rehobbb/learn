import openpyxl  # 导入 openpyxl 库，用于操作 Excel 文件
wb = openpyxl.Workbook()  # 创建一个新的工作簿对象
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
wb.create_sheet()  # 创建一个新的工作表，使用默认名称
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
wb.create_sheet(index=0,title='first sheet')  # 在索引 0 的位置创建一个名为 'first sheet' 的工作表
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
wb.create_sheet(index=2,title= 'middle sheet')  # 在索引 2 的位置创建一个名为 'middle sheet' 的工作表
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
wb.save('example3.xlsx')  # 将工作簿保存为 'example3.xlsx' 文件
del wb['middle sheet']  # 删除名为 'middle sheet' 的工作表
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
del wb['Sheet']  # 删除名为 'Sheet' 的工作表
print(wb.sheetnames)  # 打印当前工作簿中所有工作表的名称
