import openpyxl  # 导入 openpyxl 库，用于操作 Excel 文件
wb = openpyxl.load_workbook('./doc/example.xlsx')  # 加载指定路径下的 Excel 文件
sheet = wb['Sheet1']  # 获取名为 'Sheet1' 的工作表
str = sheet.cell(1,2).value  # 获取第 1 行第 2 列单元格的值
print(str)
for i in range(1,8,2):  # 遍历从 1 到 8（不包含 8），步长为 2 的整数
    print(i,sheet.cell(i,2).value)
# 获取工作表的最大行数和最大列数，并分别赋值给 marow 和 macol 变量
marow ,macol =  sheet.max_row,sheet.max_column
print(marow,macol)