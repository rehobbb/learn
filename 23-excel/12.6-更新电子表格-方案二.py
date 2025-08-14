import openpyxl  # 导入 openpyxl 库，用于操作 Excel 文件
wb = openpyxl.load_workbook('../produceSales.xlsx')  # 加载指定路径下的 Excel 文件
sheet = wb['Sheet']  # 获取名为 'Sheet' 的工作表
update_price = {  # 定义一个字典，存储需要更新的商品价格
    'Gelery': 40,  # 芹菜的新价格为 400
    'Garlic': 50,  # 大蒜的新价格为 500
    'Lemon': 60,   # 柠檬的新价格为 600
}
for row in range(2, sheet.max_row+1):  # 遍历工作表中从第 2 行开始到最后一行的每一行
    name = sheet.cell(row=row, column=1).value  # 获取当前行第一列单元格的值作为商品名称
    if name in update_price:  # 检查商品名称是否在需要更新价格的字典中
        sheet['b'+str(row)] = update_price[name]  # 如果存在，则更新当前行第二列单元格的价格
wb.save('../produceSales3.xlsx')  # 将修改后的工作簿保存到指定路径
