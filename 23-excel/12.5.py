import openpyxl  # 导入 openpyxl 库，用于操作 Excel 文件
wb = openpyxl.Workbook()  # 创建一个新的 Excel 工作簿
print(wb.sheetnames)  # 打印工作簿中所有工作表的名称
sheet = wb.active  # 获取当前活动的工作表
print(sheet.title)  # 打印当前活动工作表的标题
sheet.title = 'spam bacon eggs sheet'  # 修改当前活动工作表的标题
print(wb.sheetnames)  # 再次打印工作簿中所有工作表的名称
wb.save('example1.xlsx')  # 将工作簿保存为 example1.xlsx 文件
